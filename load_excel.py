import openpyxl

loc = r'C:\\Users\\hp\\OneDrive\\Desktop\\Data...fun.xlsx'  # Replace with your actual file path
wb = openpyxl.load_workbook(loc)

# Access the first sheet by index (0)
sheet = wb.worksheets[0]

# Access cell value at row=4 (index 3) and column=4 (index 3)
print(sheet.cell(row=4, column=4).value)

# Print the number of columns and rows
print(sheet.max_column)
print(sheet.max_row)

# Print data in the first row
for cell in sheet[1]:
    print(cell.value)

# Print data in the second row
row_values = []
for cell in sheet[2]:
    row_values.append(cell.value)
print(row_values)

# Print data in the second column
col_values = []
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
    col_values.append(row[0].value)
print(col_values)

